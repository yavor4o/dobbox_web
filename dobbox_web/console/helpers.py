import os
import googlemaps
from django.conf import settings
from dobbox_web.transport.models import TransportRequests, TransportPriceCompany, TransportCompany
import pandas as pd
from googlemaps.exceptions import ApiError
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import openpyxl
from django.core.mail import EmailMessage

def calculate_total_cost_for_requests(request_ids):
    gmaps = googlemaps.Client(key=settings.GOOGLE_MAPS_API_KEY)

    total_cost = 0
    total_distance = 0
    transport_costs = 0
    asset_costs = 0

    gps_coordinates = []
    transport_company_ids = set()
    start_end_coordinates = None

    for request_id in request_ids:
        try:
            request = TransportRequests.objects.get(pk=request_id)

            # Добавяне на ID на транспортната компания в множеството
            transport_company_ids.add(request.transport_company.pk)

            if request.is_self_transport:
                start_end_coordinates = (request.user.userdata.office.gps_lat, request.user.userdata.office.gps_long)
            else:
                start_end_coordinates = (request.transport_company.gps_lat, request.transport_company.gps_long)

            if request.contragent:
                object_coordinates = f"{request.contragent.gps_lat}, {request.contragent.gps_long}"
                gps_coordinates.append(object_coordinates)

                for asset in request.assets.all():
                    try:
                        # Вземете цената за съоръжението според типа му и асоциираната транспортна компания
                        pricing = TransportPriceCompany.objects.filter(
                            transport_company=request.transport_company.pk,
                            asset_type=asset.asset.asset_type
                        ).first()
                        asset_costs += pricing.price
                    except TransportPriceCompany.DoesNotExist:
                        # Ако няма определена цена за този тип актив и транспортна компания, може да пропуснете или да логирате грешка
                        pass

        except TransportRequests.DoesNotExist:
            continue

    if len(transport_company_ids) > 1:
        raise ValueError("Невъзможно е пускане на транспорт към различни компании.")

    if gps_coordinates and start_end_coordinates:
        directions_result = gmaps.directions(start_end_coordinates,
                                             start_end_coordinates,
                                             waypoints=gps_coordinates,
                                             optimize_waypoints=True,
                                             mode="driving")
        for leg in directions_result[0]['legs']:
            total_distance += leg['distance']['value']

    total_distance_km = total_distance / 1000

    if transport_company_ids:
        transport_company_id = next(iter(transport_company_ids))
        transport_company = TransportCompany.objects.get(id=transport_company_id)
        if transport_company:
            transport_costs = total_distance_km * float(transport_company.price_per_km)

    total_cost = float(asset_costs) + float(transport_costs)

    sorted_gps_coordinates = [{'lat': lat, 'lng': lng} for lat, lng in
                              [map(float, coord.split(',')) for coord in gps_coordinates]]

    return {
        'total_cost': total_cost,
        'total_distance_km': total_distance_km,
        'transport_costs': transport_costs,
        'asset_costs': asset_costs,
        'sorted_gps_coordinates': sorted_gps_coordinates,
        'start_end_coordinates': {'lat': start_end_coordinates[0], 'lng': start_end_coordinates[1]},
    }


def generate_excel_file(data, file_name):
    """
    Генерира Excel файл от предоставени данни и прилага специфично форматиране.

    :param data: Списък от речници, където всеки речник съдържа данни за един ред от Excel файла.
    :param file_name: Име на файла, който ще бъде създаден.
    :return: Пътят до създадения файл.
    """
    # Create DataFrame from data
    df = pd.DataFrame(data)

    # Ensure temporary directory exists
    tmp_directory = os.path.join('tmp')
    if not os.path.exists(tmp_directory):
        os.makedirs(tmp_directory)

    # Save DataFrame to Excel file in 'tmp' folder
    file_path = os.path.join('tmp', file_name)
    df.to_excel(file_path, index=False)

    # Load the workbook and get the first worksheet
    wb = load_workbook(file_path)
    ws = wb.active

    # Format the headers
    header_font = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="4F81BD")
    for cell in ws["1:1"]:
        cell.font = header_font
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")

    # Auto-adjust columns' width
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        column_letter = openpyxl.utils.get_column_letter(column_cells[0].column)
        ws.column_dimensions[column_letter].width = length

    # Set thin border for all cells
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border

    # Save the formatted workbook
    wb.save(file_path)

    return file_path


def send_email(subject, recipients, html_body, attachments=None):
    email = EmailMessage(
        subject=subject,
        body=html_body,
        from_email=settings.EMAIL_HOST_USER,
        to=recipients,
        cc=[settings.MAIL_DEFAULT_CC_RECIPIENT],  # Предполага се, че тази настройка е добавена във вашите Django settings
    )
    email.content_subtype = 'html'

    if attachments:
        for attachment in attachments:
            email.attach_file(attachment)

    email.send()
